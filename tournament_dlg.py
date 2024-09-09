from PyQt5 import QtCore, QtGui, QtWidgets
from models import PlayersModel, FilaModel, TableModel
from torneo_dialogo_ui import Ui_Form
from openpyxl import load_workbook
import gc
from rich import print

# pyright: reportIncompatibleMethodOverride=none
# pyright: reportAttributeAccessIssue=none
# pyright: reportFunctionMemberAccess=none


class TournamentDialog(QtWidgets.QWidget, Ui_Form):
    def __init__(self, parent):
        self.parent = parent
        super().__init__()
        self.setupUi(self)
        self.setWindowModality(QtCore.Qt.ApplicationModal)
        self.next_screen.clicked.connect(lambda: self.tour_navigator.setCurrentIndex(1))
        self.previous_screen.clicked.connect(
            lambda: self.tour_navigator.setCurrentIndex(0)
        )
        today = QtCore.QDate.currentDate()
        self.tour_name.setText(
            self.tour_name.text() + " " + today.toString("MMMM dd, yyyy")
        )
        self.tour_date.setDate(today)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.close)
        self.model = PlayersModel(
            self,
            [
                [
                    "ID_no",
                    "Name",
                    "ClubName",
                    "BirthDay",
                    "Sex",
                    "Rtg_nat",
                    "Rtg_int",
                    "Title",
                    "FIDE_no",
                    "Fed",
                    "ClubNo",
                    "Status",
                    "PARTIE",
                    "BODY",
                    "SUMA_ELO",
                    "k",
                ]
            ],
        )
        self.player_table.setModel(self.model)
        self.player_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.player_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.player_table.customContextMenuRequested.connect(
            self.players_table_customContextMenuRequested
        )

    def accept(self):
        self.parent.model = self.model
        self.parent.players_table.setModel(self.model)
        self.parent.navigator.setCurrentIndex(3)
        data = [{col: row[i] for i, col in enumerate(self.model.headers)} for row in self.model.players_data]
        print(data)
        self.parent.masnou_manager.create_tournament(data)
        self.parent.setup_tournament()
        self.parent.list_model = FilaModel(self.parent.masnou_manager.get_cola_names())
        self.parent.fila_model_main = FilaModel(
            self.parent.masnou_manager.get_cola_names()
        )
        self.parent.fila_list_main.setModel(self.parent.fila_model_main)
        self.parent.fila_list.setModel(self.parent.list_model)
        self.parent.score_model = TableModel(
            self.parent, self.parent.masnou_manager.return_table()
        )
        self.parent.score_table.setModel(self.parent.score_model)
        self.parent.tournament_active = True
        self.parent.tournament_title.setText(self.tour_name.text())
        self.parent.finish_button.setEnabled(True)
        self.parent.action_Editar_Jugadores.setEnabled(True)
        self.parent.action_Torneo_Actual.setEnabled(True)
        self.parent.action_Rondas.setEnabled(True)
        self.parent.action_Detalles.setEnabled(True)
        self.parent.menu_btn_go_tour.setEnabled(True)
        self.parent.menu_btn_finish.setEnabled(True)
        self.close()
        self.destroy()
        del self
        gc.collect()

    def players_table_customContextMenuRequested(self, event: QtCore.QEvent) -> None:
        menu = QtWidgets.QMenu(self)
        remove_action = menu.addAction("Remove")
        add_action = menu.addAction("Add")
        load_action = menu.addAction("Load")
        action = menu.exec_(QtGui.QCursor.pos())
        if action == remove_action:
            self.remove_player()
        elif action == add_action:
            self.add_player()
        elif action == load_action:
            self.load_players()

    def remove_player(self):
        index = self.player_table.currentIndex()
        if index == -1:
            return
        self.model.removeRows(index.row(), 1)

    def add_player(
        self,
    ):
        index = self.player_table.currentIndex()
        if index == -1:
            return
        self.model.insertRows(len(self.model.players_data), 1)

    def load_players(self):
        file, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Load Players",
            "",
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)",
        )
        if not file:
            return
        wb = load_workbook(file)
        ws = wb.active
        if ws is not None:
            self.model = PlayersModel(
                self,
                [
                    [cell.value for cell in row]
                    for row in ws.iter_rows(min_row=1, max_col=16, max_row=20)
                ],
            )
            self.player_table.setModel(self.model)
            self.player_table.resizeColumnsToContents()
