from PyQt5 import QtCore, QtGui, QtWidgets
from models import PlayersModel, FilaModel, TableModel, PairingModel
from dialogs.torneo_dialogo_ui import Ui_Form
from openpyxl import load_workbook
import gc
from rich import print

# pyright: reportIncompatibleMethodOverride=none
# pyright: reportAttributeAccessIssue=none
# pyright: reportFunctionMemberAccess=none


class TournamentDialog(QtWidgets.QWidget, Ui_Form):
    def __init__(self, parent):
        self.parent = parent
        super().__init__()
        self.setupUi(self)
        self.setWindowModality(QtCore.Qt.ApplicationModal)
        self.next_screen.clicked.connect(lambda: self.tour_navigator.setCurrentIndex(1))
        self.previous_screen.clicked.connect(
            lambda: self.tour_navigator.setCurrentIndex(0)
        )
        self.previous_screen_last.clicked.connect(
            lambda: self.tour_navigator.setCurrentIndex(1)
        )
        self.next_screen_2.clicked.connect(
            lambda: self.tour_navigator.setCurrentIndex(2)
        )
        today = QtCore.QDate.currentDate()
        self.tour_name.setText(
            self.tour_name.text() + " " + today.toString("MMMM dd, yyyy")
        )
        self.tour_date.setDate(today)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.close)
        self.pairings_model = PairingModel(self, [[0,0]])
        self.model = PlayersModel(
            self,
            [
                [
                    "ID_no",
                    "Name",
                    "ClubName",
                    "BirthDay",
                    "Sex",
                    "Rtg_nat",
                    "Rtg_int",
                    "Title",
                    "FIDE_no",
                    "Fed",
                    "ClubNo",
                    "Status",
                    "PARTIE",
                    "BODY",
                    "SUMA_ELO",
                    "k",
                ]
            ],
        )
        self.player_table.setModel(self.model)
        self.player_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.player_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.player_table.customContextMenuRequested.connect(
            self.players_table_customContextMenuRequested
        )
        # pairings table
        self.emparejamientos_table.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectRows
        )
        self.emparejamientos_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.emparejamientos_table.customContextMenuRequested.connect(
            self.pairing_table_customContextMenuRequested
        )
        self.emparejamientos_table.setModel(self.pairings_model)

    def accept(self):
        self.parent.players_model = self.model
        self.parent.players_table.setModel(self.parent.players_model)
        self.parent.navigator.setCurrentIndex(3)
        data = [
            {col: row[i] for i, col in enumerate(self.model.headers)}
            for row in self.model.players_data
        ]
        self.parent.masnou_manager.create_tournament(
            data, {"num_players_board": self.jugadores_tablero.value()}
        )
        if self.check_emparejamientos.isChecked():
            self.parent.masnou_manager.groups = dict(enumerate(self.pairings_model.pairing_data))
            ids_playing = [id for sub in self.pairings_model.pairing_data for id in sub]
            ids_all = [player["ID_no"] for player in data]
            self.parent.masnou_manager.cola = [id for id in ids_all if id not in ids_playing]
            print(dict(enumerate(self.pairings_model.pairing_data)))
        self.parent.setup_tournament()
        self.parent.list_model = FilaModel(self.parent.masnou_manager.get_cola_names())
        self.parent.fila_model_main = FilaModel(
            self.parent.masnou_manager.get_cola_names()
        )
        self.parent.fila_list_main.setModel(self.parent.fila_model_main)
        self.parent.fila_list.setModel(self.parent.list_model)
        self.parent.score_model = TableModel(
            self.parent, self.parent.masnou_manager.return_table()
        )
        self.parent.score_table.setModel(self.parent.score_model)
        self.parent.tournament_active = True
        self.parent.tournament_title.setText(self.tour_name.text())
        self.parent.finish_button.setEnabled(True)
        self.parent.action_Editar_Jugadores.setEnabled(True)
        self.parent.action_Torneo_Actual.setEnabled(True)
        self.parent.action_Rondas.setEnabled(True)
        self.parent.action_Detalles.setEnabled(True)
        self.parent.menu_btn_go_tour.setEnabled(True)
        self.parent.menu_btn_finish.setEnabled(True)
        self.close()
        self.destroy()
        del self
        gc.collect()

    def pairing_table_customContextMenuRequested(self, event):
        menu = QtWidgets.QMenu(self)
        remove_action = menu.addAction("Remove")
        add_action = menu.addAction("Añadir")
        action = menu.exec_(QtGui.QCursor.pos())
        if action == add_action:
            self.add_pairing()
        if action == remove_action:
            self.remove_pairing()

    def add_pairing(self):
        index = self.emparejamientos_table.currentIndex()
        if index == -1:
            return
        self.pairings_model.insertRows(len(self.pairings_model.pairing_data), 1)
        self.pairings_model.layoutChanged.emit()

    def remove_pairing(self):
        index = self.emparejamientos_table.currentIndex()
        if index == -1:
            return
        self.pairings_model.removeRows(index.row(), 1)
        self.pairings_model.layoutChanged.emit()

    def players_table_customContextMenuRequested(self, event: QtCore.QEvent) -> None:
        menu = QtWidgets.QMenu(self)
        remove_action = menu.addAction("Remover")
        if remove_action is not None:
            remove_action.setShortcut("Ctrl+d")
        add_action = menu.addAction("Añadir")
        load_action = menu.addAction("Importar")
        action = menu.exec_(QtGui.QCursor.pos())
        if action == remove_action:
            self.remove_player()
        elif action == add_action:
            self.add_player()
        elif action == load_action:
            self.load_players()

    def remove_player(self):
        index = self.player_table.currentIndex()
        if index == -1:
            return
        self.model.removeRows(index.row(), 1)
        self.model.layoutChanged.emit()

    def add_player(
        self,
    ):
        index = self.player_table.currentIndex()
        if index == -1:
            return
        self.model.insertRows(len(self.model.players_data), 1)
        self.model.layoutChanged.emit()

    def load_players(self):
        file, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Load Players",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file:
            return
        wb = load_workbook(file)
        ws = wb.active
        if ws is not None:
            last_row = ws.max_row
            data = [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=1, max_col=16, max_row=last_row)
            ]

            self.model = PlayersModel(self, data)
            self.model.layoutChanged.connect(
                lambda: self.update_values(self.model.players_data)
            )
            self.update_values(data)
            self.player_table.setModel(self.model)
            self.player_table.resizeColumnsToContents()

    def update_values(self, data):
        self.total_jugadores.setText(str(len(data)))
        if len(data) > 2:
            if len(data) % 2 == 0:
                self.jugadores_cola.setValue(2)
                self.jugadores_tablero.setValue(len(data) - 2)
            else:
                self.jugadores_cola.setValue(1)
                self.jugadores_tablero.setValue(len(data) - 1)
